#add this comment to check github branching.
#flaksdj lasd alñsd fañlsd

import os
import openpyxl
import time

def add_info_excel(folder_path: str, info_to_add: str):
    """Method to add some information inside specific cell inside Excel
       files. In this case we have excel files with youtube channel codes on it.
       We want to add to add to that code 'https://www.youtube.com/watch?v=', so we
       can call the channel from the excel sheet.

       Args:
           folder_path (str): String passed to the method to indicate the route where the
                              excel files are.
           info_to_add (str): String to add to the cell content.
    """

    #We put the folder we pass to the method as our current working directory
    #So we execute the .py file from wherever we want
    os.chdir(folder_path)

    excel_files = [file for file in os.listdir(folder_path) if file.endswith(('.xls', '.xlsx'))]
    for file in excel_files:
        #Open a workbook. Only interested in data not in formulas.
        wb = openpyxl.load_workbook(file, data_only=True)

        #In our case the sheet we want to modify have the same name as the excel file.
        #Change for the name of the sheet otherwise
        sheet = wb[os.path.splitext(file)[0]]

        #We go through all the rows with information. We have to begin with row number 5
        for i in range(5, sheet.max_row+1):
            #The information we want to modify is in column number 1
            if sheet.cell(row=i, column=1).value != None:
                sheet.cell(row=i, column=1).value = info_to_add.strip() + sheet.cell(row=i, column=1).value
            else:
                continue
        wb.save(file)

add_info_excel('D:/OneDrive/alexca73myCode', 'https://www.youtube.com/watch?v=')
